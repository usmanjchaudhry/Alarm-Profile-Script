
import tkinter as tk
from tkinter import filedialog
import pandas as pd
from glob import glob
import csv
import openpyxl
import numpy as np
#import seaborn as sns
from openpyxl.formatting import Rule
from openpyxl.styles import Font, PatternFill, Border
from openpyxl.styles.differential import DifferentialStyle
#from openpyxl.formatting.rule import ColorScale, FormatObject
#from openpyxl.styles import Color
#from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles.borders import Border, Side
from tkinter import Label 
from tkinter import Tk
#from tkinter import Button


#from styleframe import StyleFrame, Styler, utils




print('Starting')


#Create an instance of tkinter frame
bin = Tk()

#Set the geometry of tkinter frame
bin.geometry("750x270")

#Initialize a Label widget
Label(bin, text= "Please be patient, cleaning the data...",
font=('Helvetica 20 bold')).pack(pady=20)







files = glob('Alarm Profile for*')
files.sort()
newFiles = ' '.join(str(e) for e in files)
newFiles =newFiles.replace('Alarm Profile for','')
newFiles =newFiles.replace("'",'')
print(newFiles)


main_dataframe = pd.DataFrame(pd.read_excel(files[0]))
for i in range(1,len(files)):
    data = pd.read_excel(files[i])
    df = pd.DataFrame(data)
    main_dataframe = pd.concat([main_dataframe,df],axis=0)
    

df0 = main_dataframe.drop(columns = ['Occurrences','Terminal','Line','Location','Type','First Alarm', 'Last Alarm'])
df0 = df0.rename(columns = {'Point Name': 'newNames'})
#display(df0)




df2 = main_dataframe.drop(columns = [ 'First Alarm', 'Last Alarm'])
df2['newName'] = df2['Point Name']
df2['Point Name'] = df2['Point Name'].map(lambda x: x.rstrip('++'))
df2['Point Name'] = df2['Point Name'].map(lambda x: x.rstrip('++ 0'))
df4 = df2.astype(str).groupby('Point Name').agg(lambda x: ','.join(x))
df4['count'] = df2.groupby(['Point Name']).sum()
df4['Terminal'] = df4['Terminal'].str.split(',').str[0]
df4['Terminal'] = df4['Terminal'].map(lambda x: x.rstrip('++'))
df4['Line'] = df4['Line'].str.split(',').str[0]
df4['Line'] = df4['Line'].str.replace('---','')
df4['Location'] = df4['Location'].str.split(',').str[0]
df4['Location'] = df4['Location'].str.replace('nan','')
df4['Type'] = df4['Type'].str.split(',').str[0]
df4['Type'] = df4['Type'].str.replace('nan','')
df5 = df4
df5 = df5.rename(columns = {'Occurrences': 'trend'})
df5['newName'] = df5['newName'].str.split(',').str[0]
df5['newName'] = df5['newName'].map(lambda x: x.rstrip('++ 0'))
df5['newName'] = df5['newName'].map(lambda x: x.rstrip('++'))

df5 = df5.rename(columns = {'newName': 'Point Name'})

s = df5.trend
df5['newCol'] = ''
for i in range(0,len(df5.trend)):
    df5['newCol'][i] = [int(item) for item in s[i].split(',')]
for i in range(0,len(df5.newCol)):
    testArray = df5.newCol[i]
    testArray = np.sort(testArray)
    upperLimit = testArray.mean() + 1.82*testArray.std()
    lowerLimit = testArray.mean() - 1.82*testArray.std()
    testArray = (testArray > upperLimit) |(testArray < lowerLimit)
    df5.newCol[i] = any(testArray)

print('This is the Data Frame')
cols = df5.columns
df5 = df5[['Line','Location','Terminal','Point Name','trend','count','Type','newCol']]
df5 = df5.rename(columns = {'newCol': 'Contains Outliers'})

print(df5)
print(cols)






print('FINISHED CLEANING DATA')
#Automatically close the window after 3 seconds
bin.after(1000,lambda:bin.destroy())

bin.mainloop()


try:
    df5.to_csv('./outputFolder/ignoreThis.csv',sep=',' )
except PermissionError:
        root3 = Tk()
        root3.title("Error!")
        Label(root3, text=" Please close 'ignoreThis.csv' before continuing! ", font=('Helvetica 14 bold')).pack(pady=40)
        root3_width = 750
        root3_height = 250
        x = int(int(root3.winfo_screenwidth()/2) - int(root3_width/2))
        y = int(int(root3.winfo_screenheight()/2) - int(root3_height/2))
        root3.geometry(f"{root3_width}x{root3_height}+{x}+{y}")
        root3.mainloop()
        root3.destroy()



wb = openpyxl.Workbook()
ws = wb.active

with open('./outputFolder/ignoreThis.csv') as f:
    reader = csv.reader(f, delimiter=',')
    for row in reader:
        ws.append(row)

#for r in dataframe_to_rows(df5, index=True, header=True):
    #ws.append(r)    

#ws.delete_rows(2)
ws.delete_cols(1)

ws.auto_filter.ref= ws.dimensions 

dims = {}
for row in ws.rows:
    for cell in row:
        if cell.value:
             dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
for col, value in dims.items():
    ws.column_dimensions[col].width = value
    
ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 35
ws.column_dimensions['G'].width = 25
ws.column_dimensions['H'].width = 25


ws.column_dimensions['F'].width = 10



red_fill = PatternFill(bgColor="00FFFF00")
dxf = DifferentialStyle(fill=red_fill)
r = Rule(type="expression", dxf=dxf, stopIfTrue=True)
r.formula = ['$H1="True"']
ws.conditional_formatting.add("E1:E1000", r)


ws['I1'] = 'Dates used for data:'
ws['I2'] = newFiles


ws['A1'].font = Font(bold=True)
ws['B1'].font = Font(bold=True)
ws['C1'].font = Font(bold=True)
ws['D1'].font = Font(bold=True)
ws['E1'].font = Font(bold=True)
ws['F1'].font = Font(bold=True)
ws['G1'].font = Font(bold=True)
ws['H1'].font = Font(bold=True)
ws['I1'].font = Font(bold=True)


medium_border = Border(left=Side(style='medium'), 
                     right=Side(style='medium'), 
                     top=Side(style='medium'), 
                     bottom=Side(style='medium'))

for i in range(1,9):
    ws.cell(row=1, column=i).border = medium_border



print('finished')
try:
    wb.save('./outputFolder/file.xlsx')
except PermissionError:
        root2 = Tk()
        root2.title("Error!")
        Label(root2, text=" Please close previous 'file.xlsx' before continuing! ", font=('Helvetica 14 bold')).pack(pady=40)
        root2_width = 750
        root2_height = 250
        x = int(int(root2.winfo_screenwidth()/2) - int(root2_width/2))
        y = int(int(root2.winfo_screenheight()/2) - int(root2_height/2))
        root2.geometry(f"{root2_width}x{root2_height}+{x}+{y}")
        root2.mainloop()
        root2.destroy()
win = Tk()
win.title("Please exit when finished")
Label(win, text=" Newly created file located within 'outputFolder' under the name 'file.xlsx' ", font=('Helvetica 14 bold')).pack(pady=40)
win_width = 1000
win_height = 200
x = int(int(win.winfo_screenwidth()/2) - int(win_width/2))
y = int(int(win.winfo_screenheight()/2) - int(win_height/2))
win.geometry(f"{win_width}x{win_height}+{x}+{y}")
win.mainloop()

